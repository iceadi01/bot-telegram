import os
import csv
import logging
import sqlite3
import tempfile
from datetime import datetime, timedelta, timezone

from datetime import time as dtime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

WIB = timezone(timedelta(hours=7))

from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes

DB_PATH = os.path.join(os.path.dirname(__file__), "finance.db")

NAMA_BULAN = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]


def now_wib() -> datetime:
    return datetime.now(tz=WIB)


def format_wib(dt: datetime) -> str:
    return f"{dt.day:02d} {NAMA_BULAN[dt.month]} {dt.year} {dt.strftime('%H:%M')} WIB"


def format_tanggal(waktu_str: str) -> str:
    try:
        dt = datetime.strptime(waktu_str, "%Y-%m-%d %H:%M:%S")
        return f"{dt.day:02d} {NAMA_BULAN[dt.month]} {dt.year} {dt.strftime('%H:%M')} WIB"
    except Exception:
        return waktu_str


def format_rupiah(jumlah: float) -> str:
    return f"Rp {jumlah:,.0f}".replace(",", ".")


def format_angka(jumlah: float) -> str:
    return f"{jumlah:,.0f}".replace(",", ".")


def parse_jumlah(teks: str) -> float:
    bersih = teks.replace(".", "").replace(",", "")
    return float(bersih)


def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS transaksi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            jenis TEXT NOT NULL CHECK(jenis IN ('masuk', 'keluar')),
            jumlah REAL NOT NULL,
            kategori TEXT NOT NULL DEFAULT 'Umum',
            keterangan TEXT,
            waktu TEXT NOT NULL
        )
    """)
    try:
        c.execute("ALTER TABLE transaksi ADD COLUMN kategori TEXT NOT NULL DEFAULT 'Umum'")
    except Exception:
        pass
    c.execute("""
        CREATE TABLE IF NOT EXISTS rekap_subscribers (
            user_id INTEGER PRIMARY KEY,
            aktif   INTEGER NOT NULL DEFAULT 1
        )
    """)
    conn.commit()
    conn.close()


def get_saldo(user_id: int) -> float:
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE 0 END), 0) -
            COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0)
        FROM transaksi WHERE user_id = ?
    """, (user_id,))
    result = c.fetchone()[0]
    conn.close()
    return result


def tambah_transaksi(user_id: int, jenis: str, jumlah: float, kategori: str, keterangan: str) -> str:
    waktu = now_wib().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "INSERT INTO transaksi (user_id, jenis, jumlah, kategori, keterangan, waktu) VALUES (?, ?, ?, ?, ?, ?)",
        (user_id, jenis, jumlah, kategori, keterangan, waktu)
    )
    conn.commit()
    conn.close()
    return waktu


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nama = update.effective_user.first_name
    pesan = (
        f"👋 Halo, *{nama}*! Selamat datang di *Bot Keuangan Pribadi*.\n\n"
        "📋 *Daftar Perintah:*\n\n"
        "➕ `/m` `/masuk` `<jumlah> <kategori> <ket>` — Catat pemasukan\n"
        "➖ `/k` `/keluar` `<jumlah> <kategori> <ket>` — Catat pengeluaran\n"
        "💰 `/s` `/saldo` — Lihat saldo & ringkasan\n"
        "📜 `/r` `/riwayat` — 10 transaksi terakhir\n"
        "📊 `/l` `/laporan` — Laporan per periode\n"
        "🏷 `/lk` `/laporan_kategori` — Laporan per kategori\n"
        "📁 `/ex` `/export` — Download semua data (CSV)\n"
        "🗑 `/h` `/hapus` `<id>` — Hapus transaksi\n"
        "✏️ `/e` `/edit` `<id> <jml> <kat> <ket>` — Edit transaksi\n"
        "💳 `/editsaldo` `<jumlah>` — Sesuaikan saldo\n"
        "🔄 `/reset` — Hapus semua data\n\n"
        "📌 *Format input:*\n"
        "`/m <jumlah> <kategori> <keterangan>`\n\n"
        "_Contoh:_\n"
        "`/m 500.000 Gaji Freelance bulan Juli`\n"
        "`/k 50.000 Makan Makan siang warteg`\n"
        "`/k 20.000 Transport Ojek ke kantor`\n"
        "`/e 3 75.000 Makan Makan malam restoran`\n\n"
        "💡 Kategori = 1 kata (Gaji, Makan, Transport, dll)\n"
        "💡 Nominal boleh pakai titik: `500.000` atau `1.500.000`"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def masuk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args = context.args

    if len(args) < 2:
        await update.message.reply_text(
            "⚠️ Format salah. Gunakan:\n`/m <jumlah> <kategori> <keterangan>`\n\n"
            "_Contoh: `/m 500.000 Gaji Freelance bulan Juli`_\n"
            "_Atau: `/m 500.000 Gaji harian`_ (kategori otomatis: Umum)",
            parse_mode="Markdown"
        )
        return

    try:
        jumlah = parse_jumlah(args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Jumlah tidak valid. Contoh: `500.000` atau `500000`", parse_mode="Markdown")
        return

    if jumlah <= 0:
        await update.message.reply_text("⚠️ Jumlah harus lebih dari 0.", parse_mode="Markdown")
        return

    if len(args) >= 3:
        kategori = args[1].capitalize()
        keterangan = " ".join(args[2:])
    else:
        kategori = "Umum"
        keterangan = args[1]

    waktu = tambah_transaksi(user_id, "masuk", jumlah, kategori, keterangan)
    saldo = get_saldo(user_id)

    pesan = (
        f"✅ *Pemasukan dicatat!*\n\n"
        f"📥 Jumlah: *{format_rupiah(jumlah)}*\n"
        f"🏷 Kategori: {kategori}\n"
        f"📝 Keterangan: {keterangan}\n"
        f"🕐 Waktu: {format_tanggal(waktu)}\n"
        f"💰 Saldo sekarang: *{format_rupiah(saldo)}*"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def keluar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args = context.args

    if len(args) < 2:
        await update.message.reply_text(
            "⚠️ Format salah. Gunakan:\n`/k <jumlah> <kategori> <keterangan>`\n\n"
            "_Contoh: `/k 50.000 Makan Makan siang warteg`_\n"
            "_Atau: `/k 50.000 Makan siang`_ (kategori otomatis: Umum)",
            parse_mode="Markdown"
        )
        return

    try:
        jumlah = parse_jumlah(args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Jumlah tidak valid. Contoh: `50.000` atau `50000`", parse_mode="Markdown")
        return

    if jumlah <= 0:
        await update.message.reply_text("⚠️ Jumlah harus lebih dari 0.", parse_mode="Markdown")
        return

    saldo_sekarang = get_saldo(user_id)
    if jumlah > saldo_sekarang:
        await update.message.reply_text(
            f"⚠️ Saldo tidak mencukupi.\n"
            f"💰 Saldo saat ini: *{format_rupiah(saldo_sekarang)}*",
            parse_mode="Markdown"
        )
        return

    if len(args) >= 3:
        kategori = args[1].capitalize()
        keterangan = " ".join(args[2:])
    else:
        kategori = "Umum"
        keterangan = args[1]

    waktu = tambah_transaksi(user_id, "keluar", jumlah, kategori, keterangan)
    saldo = get_saldo(user_id)

    pesan = (
        f"✅ *Pengeluaran dicatat!*\n\n"
        f"📤 Jumlah: *{format_rupiah(jumlah)}*\n"
        f"🏷 Kategori: {kategori}\n"
        f"📝 Keterangan: {keterangan}\n"
        f"🕐 Waktu: {format_tanggal(waktu)}\n"
        f"💰 Saldo sekarang: *{format_rupiah(saldo)}*"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def saldo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0),
            COUNT(*)
        FROM transaksi WHERE user_id = ?
    """, (user_id,))
    row = c.fetchone()
    conn.close()

    total_masuk, total_keluar, total_transaksi = row
    saldo_bersih = total_masuk - total_keluar
    now = now_wib()

    pesan = (
        f"💰 *Ringkasan Keuangan Anda*\n"
        f"🕐 {format_wib(now)}\n\n"
        f"📥 Total Pemasukan: *{format_rupiah(total_masuk)}*\n"
        f"📤 Total Pengeluaran: *{format_rupiah(total_keluar)}*\n"
        f"🔢 Total Transaksi: *{total_transaksi}x*\n"
        f"─────────────────\n"
        f"💵 Saldo Bersih: *{format_rupiah(saldo_bersih)}*"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def riwayat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    args = context.args
    limit = 10
    if args:
        try:
            limit = max(1, min(int(args[0]), 50))
        except ValueError:
            pass

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT id, jenis, jumlah, kategori, keterangan, waktu
        FROM transaksi
        WHERE user_id = ?
        ORDER BY waktu DESC
        LIMIT ?
    """, (user_id, limit))
    rows = c.fetchall()
    conn.close()

    if not rows:
        await update.message.reply_text("📭 Belum ada transaksi tercatat.", parse_mode="Markdown")
        return

    lines = [f"📜 *{limit} Transaksi Terakhir:*\n"]
    for row in rows:
        tid, jenis, jumlah, kategori, keterangan, waktu = row
        ikon = "📥" if jenis == "masuk" else "📤"
        tanda = "+" if jenis == "masuk" else "-"
        lines.append(
            f"{ikon} `#{tid}` {format_tanggal(waktu)}\n"
            f"   {tanda}{format_rupiah(jumlah)} | 🏷{kategori} | {keterangan}"
        )

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def laporan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    now = now_wib()

    hari_ini = now.strftime("%Y-%m-%d")
    awal_minggu = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    awal_bulan = now.strftime("%Y-%m-01")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    def ambil_periode(sejak: str):
        c.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0),
                COUNT(*)
            FROM transaksi
            WHERE user_id = ? AND DATE(waktu) >= ?
        """, (user_id, sejak))
        return c.fetchone()

    h_masuk, h_keluar, h_jml = ambil_periode(hari_ini)
    m_masuk, m_keluar, m_jml = ambil_periode(awal_minggu)
    b_masuk, b_keluar, b_jml = ambil_periode(awal_bulan)

    conn.close()

    tgl_awal_minggu = (now - timedelta(days=now.weekday()))

    pesan = (
        f"📊 *Laporan Keuangan*\n\n"
        f"📅 *Hari Ini* ({now.day:02d} {NAMA_BULAN[now.month]} {now.year})\n"
        f"  📥 Masuk: {format_rupiah(h_masuk)}\n"
        f"  📤 Keluar: {format_rupiah(h_keluar)}\n"
        f"  💵 Selisih: {format_rupiah(h_masuk - h_keluar)}  ({h_jml} transaksi)\n\n"
        f"📅 *Minggu Ini* (sejak {tgl_awal_minggu.day:02d} {NAMA_BULAN[tgl_awal_minggu.month]})\n"
        f"  📥 Masuk: {format_rupiah(m_masuk)}\n"
        f"  📤 Keluar: {format_rupiah(m_keluar)}\n"
        f"  💵 Selisih: {format_rupiah(m_masuk - m_keluar)}  ({m_jml} transaksi)\n\n"
        f"📅 *Bulan Ini* ({NAMA_BULAN[now.month]} {now.year})\n"
        f"  📥 Masuk: {format_rupiah(b_masuk)}\n"
        f"  📤 Keluar: {format_rupiah(b_keluar)}\n"
        f"  💵 Selisih: {format_rupiah(b_masuk - b_keluar)}  ({b_jml} transaksi)\n\n"
        f"_Gunakan /lk untuk laporan per kategori_"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def laporan_kategori(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    now = now_wib()
    awal_bulan = now.strftime("%Y-%m-01")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("""
        SELECT kategori,
               COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE 0 END), 0) AS total_masuk,
               COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0) AS total_keluar,
               COUNT(*) AS jml
        FROM transaksi
        WHERE user_id = ? AND DATE(waktu) >= ?
        GROUP BY kategori
        ORDER BY total_keluar DESC
    """, (user_id, awal_bulan))
    rows = c.fetchall()
    conn.close()

    if not rows:
        await update.message.reply_text(
            f"📭 Belum ada transaksi di {NAMA_BULAN[now.month]} {now.year}.",
            parse_mode="Markdown"
        )
        return

    lines = [f"🏷 *Laporan Kategori — {NAMA_BULAN[now.month]} {now.year}*\n"]
    for kat, t_masuk, t_keluar, jml in rows:
        selisih = t_masuk - t_keluar
        tanda = "▲" if selisih >= 0 else "▼"
        lines.append(
            f"📌 *{kat}* ({jml}x)\n"
            f"   📥 {format_rupiah(t_masuk)}  📤 {format_rupiah(t_keluar)}\n"
            f"   {tanda} Selisih: {format_rupiah(selisih)}"
        )

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


def buat_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def buat_excel(rows, ringkasan_kat):
    wb = Workbook()

    # ── Warna ──────────────────────────────────────────────────
    HIJAU_TUA   = "1A7A4A"
    HIJAU_MID   = "2ECC71"
    HIJAU_MUDA  = "D5F5E3"
    MERAH_TUA   = "C0392B"
    MERAH_MUDA  = "FADBD8"
    ABU_HEADER  = "2C3E50"
    ABU_MUDA    = "F2F3F4"
    PUTIH       = "FFFFFF"
    KUNING      = "F9E79F"

    fill_header   = PatternFill("solid", fgColor=ABU_HEADER)
    fill_masuk    = PatternFill("solid", fgColor=HIJAU_MUDA)
    fill_keluar   = PatternFill("solid", fgColor=MERAH_MUDA)
    fill_alt      = PatternFill("solid", fgColor=ABU_MUDA)
    fill_total    = PatternFill("solid", fgColor=KUNING)

    font_header   = Font(bold=True, color=PUTIH, size=11)
    font_bold     = Font(bold=True, size=10)
    font_normal   = Font(size=10)
    font_total    = Font(bold=True, size=11)

    tengah = Alignment(horizontal="center", vertical="center")
    kiri   = Alignment(horizontal="left",   vertical="center")
    kanan  = Alignment(horizontal="right",  vertical="center")

    border = buat_border()

    # ════════════════════════════════════════════════════════════
    # SHEET 1 — DATA TRANSAKSI
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Transaksi"
    ws.sheet_view.showGridLines = False

    # Judul
    ws.merge_cells("A1:G1")
    ws["A1"] = "📒 LAPORAN KEUANGAN PRIBADI"
    ws["A1"].font = Font(bold=True, size=14, color=ABU_HEADER)
    ws["A1"].alignment = tengah
    ws["A1"].fill = PatternFill("solid", fgColor=HIJAU_MUDA)

    ws.merge_cells("A2:G2")
    now = now_wib()
    ws["A2"] = f"Diekspor: {now.day:02d} {NAMA_BULAN[now.month]} {now.year} {now.strftime('%H:%M')} WIB  |  {len(rows)} transaksi"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = tengah

    # Header tabel
    headers = ["No", "Tanggal", "Waktu (WIB)", "Jenis", "Kategori", "Jumlah (Rp)", "Keterangan"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = tengah
        cell.border = border

    # Lebar kolom
    lebar = [5, 18, 13, 13, 13, 16, 35]
    for i, w in enumerate(lebar, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 22

    # Data baris
    total_masuk = 0
    total_keluar = 0
    for i, row in enumerate(rows, start=1):
        tid, jenis, jumlah, kategori, keterangan, waktu = row
        try:
            dt = datetime.strptime(waktu, "%Y-%m-%d %H:%M:%S")
            tgl = f"{dt.day:02d} {NAMA_BULAN[dt.month]} {dt.year}"
            jam = dt.strftime("%H:%M") + " WIB"
        except Exception:
            tgl, jam = waktu, ""

        r = i + 4
        is_masuk = jenis == "masuk"
        fill_row = fill_masuk if is_masuk else fill_keluar

        data = [i, tgl, jam, "Pemasukan" if is_masuk else "Pengeluaran",
                kategori, jumlah, keterangan]
        aligns = [tengah, tengah, tengah, tengah, tengah, kanan, kiri]

        for col, (val, aln) in enumerate(zip(data, aligns), start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill_row if col <= 6 else (fill_masuk if is_masuk else fill_keluar)
            cell.alignment = aln
            cell.border = border
            cell.font = font_normal

        # Format angka kolom F
        ws.cell(row=r, column=6).number_format = '#,##0'

        if is_masuk:
            total_masuk += jumlah
        else:
            total_keluar += jumlah

    # Baris total
    r_total = len(rows) + 5
    ws.merge_cells(f"A{r_total}:E{r_total}")
    ws[f"A{r_total}"] = "TOTAL"
    ws[f"A{r_total}"].font = font_total
    ws[f"A{r_total}"].fill = fill_total
    ws[f"A{r_total}"].alignment = tengah
    ws[f"A{r_total}"].border = border

    ws[f"F{r_total}"] = total_masuk - total_keluar
    ws[f"F{r_total}"].font = font_total
    ws[f"F{r_total}"].fill = fill_total
    ws[f"F{r_total}"].alignment = kanan
    ws[f"F{r_total}"].border = border
    ws[f"F{r_total}"].number_format = '#,##0'

    ws[f"G{r_total}"] = f"Masuk: Rp {total_masuk:,.0f}  |  Keluar: Rp {total_keluar:,.0f}".replace(",", ".")
    ws[f"G{r_total}"].font = font_total
    ws[f"G{r_total}"].fill = fill_total
    ws[f"G{r_total}"].alignment = kiri
    ws[f"G{r_total}"].border = border

    # ════════════════════════════════════════════════════════════
    # SHEET 2 — RINGKASAN & GRAFIK
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Ringkasan & Grafik")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18

    # Judul
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "📊 RINGKASAN KEUANGAN"
    ws2["A1"].font = Font(bold=True, size=14, color=ABU_HEADER)
    ws2["A1"].alignment = tengah
    ws2["A1"].fill = PatternFill("solid", fgColor=HIJAU_MUDA)
    ws2.row_dimensions[1].height = 26

    # Tabel ringkasan pemasukan/pengeluaran
    headers2 = ["Keterangan", "Jumlah (Rp)", "Persentase"]
    for col, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = tengah
        cell.border = border

    saldo_bersih = total_masuk - total_keluar
    total_all = total_masuk + total_keluar if (total_masuk + total_keluar) > 0 else 1
    pct_masuk  = total_masuk  / total_all * 100
    pct_keluar = total_keluar / total_all * 100

    ringkasan_data = [
        ("💚 Total Pemasukan",  total_masuk,   pct_masuk,  fill_masuk),
        ("❤️ Total Pengeluaran", total_keluar, pct_keluar, fill_keluar),
        ("💛 Saldo Bersih",      saldo_bersih,  None,       fill_total),
    ]
    for r_i, (label, nilai, pct, fill) in enumerate(ringkasan_data, start=4):
        ws2.cell(row=r_i, column=1, value=label).fill = fill
        ws2.cell(row=r_i, column=1).font = font_bold
        ws2.cell(row=r_i, column=1).border = border
        ws2.cell(row=r_i, column=1).alignment = kiri

        ws2.cell(row=r_i, column=2, value=nilai).fill = fill
        ws2.cell(row=r_i, column=2).font = font_bold
        ws2.cell(row=r_i, column=2).border = border
        ws2.cell(row=r_i, column=2).alignment = kanan
        ws2.cell(row=r_i, column=2).number_format = '#,##0'

        pct_val = f"{pct:.1f}%" if pct is not None else "-"
        ws2.cell(row=r_i, column=3, value=pct_val).fill = fill
        ws2.cell(row=r_i, column=3).font = font_bold
        ws2.cell(row=r_i, column=3).border = border
        ws2.cell(row=r_i, column=3).alignment = tengah

    # ── Grafik Pie: Pemasukan vs Pengeluaran ──────────────────
    pie_data_row = 10
    ws2.cell(row=pie_data_row,     column=1, value="Pemasukan")
    ws2.cell(row=pie_data_row,     column=2, value=total_masuk)
    ws2.cell(row=pie_data_row + 1, column=1, value="Pengeluaran")
    ws2.cell(row=pie_data_row + 1, column=2, value=total_keluar)

    pie = PieChart()
    pie.title = "Pemasukan vs Pengeluaran"
    pie.style = 10
    pie.width = 14
    pie.height = 10

    data_ref  = Reference(ws2, min_col=2, min_row=pie_data_row, max_row=pie_data_row + 1)
    label_ref = Reference(ws2, min_col=1, min_row=pie_data_row, max_row=pie_data_row + 1)
    pie.add_data(data_ref)
    pie.set_categories(label_ref)
    pie.series[0].graphicalProperties.solidFill = HIJAU_MID

    slice_hijau = DataPoint(idx=0)
    slice_hijau.graphicalProperties.solidFill = HIJAU_MID
    slice_merah = DataPoint(idx=1)
    slice_merah.graphicalProperties.solidFill = "E74C3C"
    pie.series[0].dPt = [slice_hijau, slice_merah]
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False

    ws2.add_chart(pie, "E2")

    # ── Grafik Bar: Pengeluaran per Kategori ─────────────────
    if ringkasan_kat:
        bar_start = pie_data_row + 4
        ws2.cell(row=bar_start - 1, column=1, value="Pengeluaran per Kategori")
        ws2.cell(row=bar_start - 1, column=1).font = Font(bold=True, size=11, color=ABU_HEADER)

        for r_i, (kat, _, t_keluar, _jml) in enumerate(ringkasan_kat, start=bar_start):
            ws2.cell(row=r_i, column=1, value=kat)
            ws2.cell(row=r_i, column=2, value=t_keluar)

        bar = BarChart()
        bar.type   = "col"
        bar.style  = 10
        bar.title  = "Pengeluaran per Kategori"
        bar.y_axis.title = "Jumlah (Rp)"
        bar.x_axis.title = "Kategori"
        bar.width  = 18
        bar.height = 12
        bar.grouping = "clustered"

        data_ref2  = Reference(ws2, min_col=2, min_row=bar_start, max_row=bar_start + len(ringkasan_kat) - 1)
        label_ref2 = Reference(ws2, min_col=1, min_row=bar_start, max_row=bar_start + len(ringkasan_kat) - 1)
        bar.add_data(data_ref2)
        bar.set_categories(label_ref2)
        bar.series[0].graphicalProperties.solidFill = MERAH_TUA

        ws2.add_chart(bar, "E20")

    return wb


async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT id, jenis, jumlah, kategori, keterangan, waktu
        FROM transaksi WHERE user_id = ?
        ORDER BY waktu ASC
    """, (user_id,))
    rows = c.fetchall()

    c.execute("""
        SELECT kategori,
               COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE 0 END), 0),
               COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0),
               COUNT(*)
        FROM transaksi
        WHERE user_id = ?
        GROUP BY kategori
        ORDER BY SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END) DESC
    """, (user_id,))
    ringkasan_kat = c.fetchall()
    conn.close()

    if not rows:
        await update.message.reply_text("📭 Belum ada transaksi untuk diekspor.", parse_mode="Markdown")
        return

    await update.message.reply_text("⏳ Sedang membuat file Excel...", parse_mode="Markdown")

    wb = buat_excel(rows, ringkasan_kat)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path = f.name
    wb.save(tmp_path)

    nama_file = f"keuangan_{now_wib().strftime('%d-%m-%Y_%H%M')}.xlsx"

    total_masuk  = sum(r[2] for r in rows if r[1] == "masuk")
    total_keluar = sum(r[2] for r in rows if r[1] == "keluar")
    saldo        = total_masuk - total_keluar

    with open(tmp_path, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=nama_file,
            caption=(
                f"📊 *Laporan Keuangan Excel*\n\n"
                f"📅 {format_wib(now_wib())}\n"
                f"🔢 {len(rows)} transaksi\n\n"
                f"📥 Pemasukan: *{format_rupiah(total_masuk)}*\n"
                f"📤 Pengeluaran: *{format_rupiah(total_keluar)}*\n"
                f"💰 Saldo: *{format_rupiah(saldo)}*\n\n"
                f"_File berisi 2 sheet: Transaksi & Ringkasan+Grafik_"
            ),
            parse_mode="Markdown"
        )

    os.unlink(tmp_path)


async def hapus(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args = context.args

    if not args:
        await update.message.reply_text(
            "⚠️ Masukkan ID transaksi.\nGunakan: `/h <id>`\n\n"
            "_Lihat ID di `/r`_",
            parse_mode="Markdown"
        )
        return

    try:
        tid = int(args[0].lstrip("#"))
    except ValueError:
        await update.message.reply_text("⚠️ ID tidak valid. Contoh: `/h 5`", parse_mode="Markdown")
        return

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT jenis, jumlah, kategori, keterangan, waktu FROM transaksi WHERE id = ? AND user_id = ?",
        (tid, user_id)
    )
    row = c.fetchone()

    if not row:
        conn.close()
        await update.message.reply_text(
            f"⚠️ Transaksi `#{tid}` tidak ditemukan atau bukan milik Anda.",
            parse_mode="Markdown"
        )
        return

    jenis, jumlah, kategori, keterangan, waktu = row
    c.execute("DELETE FROM transaksi WHERE id = ? AND user_id = ?", (tid, user_id))
    conn.commit()
    conn.close()

    saldo = get_saldo(user_id)
    ikon = "📥" if jenis == "masuk" else "📤"

    pesan = (
        f"🗑 *Transaksi dihapus!*\n\n"
        f"{ikon} `#{tid}` — {format_rupiah(jumlah)}\n"
        f"🏷 Kategori: {kategori}\n"
        f"📝 Keterangan: {keterangan}\n"
        f"🕐 Waktu: {format_tanggal(waktu)}\n"
        f"─────────────────\n"
        f"💰 Saldo sekarang: *{format_rupiah(saldo)}*"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def edit_transaksi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args = context.args

    if len(args) < 4:
        await update.message.reply_text(
            "⚠️ Format salah. Gunakan:\n`/e <id> <jumlah> <kategori> <keterangan>`\n\n"
            "_Contoh: `/e 3 75.000 Makan Makan malam restoran`_\n\n"
            "_Lihat ID di /r_",
            parse_mode="Markdown"
        )
        return

    try:
        tid = int(args[0].lstrip("#"))
    except ValueError:
        await update.message.reply_text("⚠️ ID tidak valid. Contoh: `/e 3 75.000 Makan Makan malam`", parse_mode="Markdown")
        return

    try:
        jumlah_baru = parse_jumlah(args[1])
    except ValueError:
        await update.message.reply_text("⚠️ Jumlah tidak valid. Contoh: `75.000` atau `75000`", parse_mode="Markdown")
        return

    if jumlah_baru <= 0:
        await update.message.reply_text("⚠️ Jumlah harus lebih dari 0.", parse_mode="Markdown")
        return

    kategori_baru = args[2].capitalize()
    keterangan_baru = " ".join(args[3:])

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT jenis, jumlah, kategori, keterangan, waktu FROM transaksi WHERE id = ? AND user_id = ?",
        (tid, user_id)
    )
    row = c.fetchone()

    if not row:
        conn.close()
        await update.message.reply_text(
            f"⚠️ Transaksi `#{tid}` tidak ditemukan atau bukan milik Anda.",
            parse_mode="Markdown"
        )
        return

    jenis_lama, jumlah_lama, kategori_lama, keterangan_lama, waktu_lama = row

    c.execute(
        "UPDATE transaksi SET jumlah = ?, kategori = ?, keterangan = ? WHERE id = ? AND user_id = ?",
        (jumlah_baru, kategori_baru, keterangan_baru, tid, user_id)
    )
    conn.commit()
    conn.close()

    saldo = get_saldo(user_id)
    ikon = "📥" if jenis_lama == "masuk" else "📤"

    pesan = (
        f"✏️ *Transaksi diperbarui!*\n\n"
        f"{ikon} ID: `#{tid}` ({jenis_lama.capitalize()})\n\n"
        f"*Sebelum:*\n"
        f"  💵 {format_rupiah(jumlah_lama)} | 🏷 {kategori_lama}\n"
        f"  📝 {keterangan_lama}\n\n"
        f"*Sesudah:*\n"
        f"  💵 {format_rupiah(jumlah_baru)} | 🏷 {kategori_baru}\n"
        f"  📝 {keterangan_baru}\n\n"
        f"🕐 Dicatat: {format_tanggal(waktu_lama)}\n"
        f"─────────────────\n"
        f"💰 Saldo sekarang: *{format_rupiah(saldo)}*"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def editsaldo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args = context.args

    if not args:
        await update.message.reply_text(
            "⚠️ Masukkan nominal saldo baru.\nGunakan: `/editsaldo <jumlah>`\n\n"
            "_Contoh: `/editsaldo 1.500.000`_",
            parse_mode="Markdown"
        )
        return

    try:
        saldo_baru = parse_jumlah(args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Jumlah tidak valid. Contoh: `/editsaldo 1.500.000`", parse_mode="Markdown")
        return

    if saldo_baru < 0:
        await update.message.reply_text("⚠️ Saldo tidak boleh minus.", parse_mode="Markdown")
        return

    saldo_lama = get_saldo(user_id)
    selisih = saldo_baru - saldo_lama

    if selisih == 0:
        await update.message.reply_text(
            f"ℹ️ Saldo sudah *{format_rupiah(saldo_lama)}*, tidak ada perubahan.",
            parse_mode="Markdown"
        )
        return

    if selisih > 0:
        waktu = tambah_transaksi(user_id, "masuk", selisih, "Penyesuaian", "✏️ Penyesuaian saldo")
        arah = f"+{format_rupiah(selisih)}"
    else:
        waktu = tambah_transaksi(user_id, "keluar", abs(selisih), "Penyesuaian", "✏️ Penyesuaian saldo")
        arah = f"-{format_rupiah(abs(selisih))}"

    pesan = (
        f"✅ *Saldo berhasil disesuaikan!*\n\n"
        f"📊 Saldo lama: {format_rupiah(saldo_lama)}\n"
        f"🔧 Penyesuaian: {arah}\n"
        f"🕐 Waktu: {format_tanggal(waktu)}\n"
        f"─────────────────\n"
        f"💰 Saldo sekarang: *{format_rupiah(saldo_baru)}*"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args = context.args

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM transaksi WHERE user_id = ?", (user_id,))
    jumlah_transaksi = c.fetchone()[0]
    conn.close()

    if not args or args[0].upper() != "YA":
        pesan = (
            f"⚠️ *Peringatan!*\n\n"
            f"Perintah ini akan menghapus *semua {jumlah_transaksi} transaksi* Anda secara permanen.\n"
            f"Saldo akan kembali ke *Rp 0*.\n\n"
            f"Jika yakin, ketik:\n`/reset YA`"
        )
        await update.message.reply_text(pesan, parse_mode="Markdown")
        return

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM transaksi WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

    pesan = (
        f"🔄 *Reset selesai!*\n\n"
        f"🗑 {jumlah_transaksi} transaksi telah dihapus.\n"
        f"💰 Saldo sekarang: *Rp 0*\n\n"
        f"_Gunakan /m untuk mulai mencatat kembali._"
    )
    await update.message.reply_text(pesan, parse_mode="Markdown")


async def kirim_rekap_bulanan(context):
    now = now_wib()
    bulan = now.month - 1 if now.month > 1 else 12
    tahun = now.year if now.month > 1 else now.year - 1
    nama_bln = NAMA_BULAN[bulan]

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT user_id FROM rekap_subscribers WHERE aktif = 1")
    users = c.fetchall()
    conn.close()

    for (uid,) in users:
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        c2.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN jenis='masuk'  THEN jumlah ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0),
                COUNT(*)
            FROM transaksi
            WHERE user_id = ?
              AND strftime('%Y', waktu) = ?
              AND strftime('%m', waktu) = ?
        """, (uid, str(tahun), f"{bulan:02d}"))
        row = c2.fetchone()
        c2.execute("""
            SELECT kategori,
                   COALESCE(SUM(CASE WHEN jenis='masuk'  THEN jumlah ELSE 0 END), 0),
                   COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0)
            FROM transaksi
            WHERE user_id = ?
              AND strftime('%Y', waktu) = ?
              AND strftime('%m', waktu) = ?
            GROUP BY kategori
            ORDER BY SUM(jumlah) DESC
        """, (uid, str(tahun), f"{bulan:02d}"))
        kat_rows = c2.fetchall()
        conn2.close()

        if not row or row[2] == 0:
            continue

        total_masuk, total_keluar, _ = row
        saldo_bersih = total_masuk - total_keluar
        saldo_total  = get_saldo(uid)

        lines = [
            f"📅 *Rekap Otomatis — {nama_bln} {tahun}*\n",
            f"🔢 Total Transaksi: {row[2]}",
            f"📥 Pemasukan:    *{format_rupiah(total_masuk)}*",
            f"📤 Pengeluaran:  *{format_rupiah(total_keluar)}*",
            f"─────────────────",
            f"💵 Surplus/Defisit: *{format_rupiah(saldo_bersih)}*",
            f"💰 Saldo Total:     *{format_rupiah(saldo_total)}*",
        ]
        if kat_rows:
            lines.append("\n📊 *Per Kategori:*")
            for kat, tm, tk in kat_rows:
                if tm > 0 and tk > 0:
                    lines.append(f"  • {kat}: 📥{format_rupiah(tm)}  📤{format_rupiah(tk)}")
                elif tm > 0:
                    lines.append(f"  • {kat}: 📥 {format_rupiah(tm)}")
                else:
                    lines.append(f"  • {kat}: 📤 {format_rupiah(tk)}")

        try:
            await context.bot.send_message(chat_id=uid, text="\n".join(lines), parse_mode="Markdown")
            logger.info(f"Rekap bulanan dikirim ke user {uid}")
        except Exception as e:
            logger.error(f"Gagal kirim rekap ke {uid}: {e}")


async def rekap(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args = context.args

    if args and args[0].lower() == "on":
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "INSERT OR REPLACE INTO rekap_subscribers (user_id, aktif) VALUES (?, 1)",
            (user_id,)
        )
        conn.commit()
        conn.close()
        await update.message.reply_text(
            "✅ *Rekap otomatis aktif!*\n\n"
            "Setiap tanggal 1 jam 07:00 WIB kamu akan menerima ringkasan bulan sebelumnya secara otomatis.\n\n"
            "Ketik `/rekap off` untuk menonaktifkan.\n"
            "Ketik `/rekap` kapan saja untuk melihat rekap bulan ini.",
            parse_mode="Markdown"
        )
        return

    if args and args[0].lower() == "off":
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("UPDATE rekap_subscribers SET aktif = 0 WHERE user_id = ?", (user_id,))
        conn.commit()
        conn.close()
        await update.message.reply_text("⛔ Rekap otomatis dinonaktifkan.", parse_mode="Markdown")
        return

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT aktif FROM rekap_subscribers WHERE user_id = ?", (user_id,))
    sub = c.fetchone()
    conn.close()

    now   = now_wib()
    bulan = now.month
    tahun = now.year

    conn2 = sqlite3.connect(DB_PATH)
    c2 = conn2.cursor()
    c2.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN jenis='masuk'  THEN jumlah ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0),
            COUNT(*)
        FROM transaksi
        WHERE user_id = ?
          AND strftime('%Y', waktu) = ?
          AND strftime('%m', waktu) = ?
    """, (user_id, str(tahun), f"{bulan:02d}"))
    row = c2.fetchone()
    c2.execute("""
        SELECT kategori,
               COALESCE(SUM(CASE WHEN jenis='masuk'  THEN jumlah ELSE 0 END), 0),
               COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END), 0)
        FROM transaksi
        WHERE user_id = ?
          AND strftime('%Y', waktu) = ?
          AND strftime('%m', waktu) = ?
        GROUP BY kategori
        ORDER BY SUM(jumlah) DESC
    """, (user_id, str(tahun), f"{bulan:02d}"))
    kat_rows = c2.fetchall()
    conn2.close()

    if not row or row[2] == 0:
        await update.message.reply_text(
            f"📭 Belum ada transaksi di bulan {NAMA_BULAN[bulan]} {tahun}.\n\n"
            "_Gunakan /rekap on untuk aktifkan rekap otomatis tiap tanggal 1._",
            parse_mode="Markdown"
        )
        return

    total_masuk, total_keluar, jml_transaksi = row
    saldo_bersih = total_masuk - total_keluar
    saldo_total  = get_saldo(user_id)

    lines = [
        f"📅 *Rekap {NAMA_BULAN[bulan]} {tahun}*\n",
        f"🔢 Total Transaksi: {jml_transaksi}",
        f"📥 Pemasukan:    *{format_rupiah(total_masuk)}*",
        f"📤 Pengeluaran:  *{format_rupiah(total_keluar)}*",
        f"─────────────────",
        f"💵 Surplus/Defisit: *{format_rupiah(saldo_bersih)}*",
        f"💰 Saldo Total:     *{format_rupiah(saldo_total)}*",
    ]

    if kat_rows:
        lines.append("\n📊 *Per Kategori:*")
        for kat, tm, tk in kat_rows:
            if tm > 0 and tk > 0:
                lines.append(f"  • {kat}: 📥{format_rupiah(tm)}  📤{format_rupiah(tk)}")
            elif tm > 0:
                lines.append(f"  • {kat}: 📥 {format_rupiah(tm)}")
            else:
                lines.append(f"  • {kat}: 📤 {format_rupiah(tk)}")

    if not sub or sub[0] == 0:
        lines.append("\n💡 _Ketik /rekap on untuk rekap otomatis tiap tanggal 1 jam 07:00 WIB._")
    else:
        lines.append("\n✅ _Rekap otomatis aktif — kamu akan dikirimi tiap tanggal 1._")

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


def main():
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise ValueError("TELEGRAM_BOT_TOKEN tidak ditemukan di environment variables.")

    init_db()
    logger.info("Database diinisialisasi.")

    app = ApplicationBuilder().token(token).build()

    app.add_handler(CommandHandler("start", start))

    app.add_handler(CommandHandler("masuk", masuk))
    app.add_handler(CommandHandler("m", masuk))

    app.add_handler(CommandHandler("keluar", keluar))
    app.add_handler(CommandHandler("k", keluar))

    app.add_handler(CommandHandler("saldo", saldo))
    app.add_handler(CommandHandler("s", saldo))

    app.add_handler(CommandHandler("riwayat", riwayat))
    app.add_handler(CommandHandler("r", riwayat))

    app.add_handler(CommandHandler("laporan", laporan))
    app.add_handler(CommandHandler("l", laporan))

    app.add_handler(CommandHandler("laporan_kategori", laporan_kategori))
    app.add_handler(CommandHandler("lk", laporan_kategori))

    app.add_handler(CommandHandler("export", export))
    app.add_handler(CommandHandler("ex", export))

    app.add_handler(CommandHandler("hapus", hapus))
    app.add_handler(CommandHandler("h", hapus))

    app.add_handler(CommandHandler("edit", edit_transaksi))
    app.add_handler(CommandHandler("e", edit_transaksi))

    app.add_handler(CommandHandler("editsaldo", editsaldo))
    app.add_handler(CommandHandler("reset", reset))

    app.add_handler(CommandHandler("rekap", rekap))

    if app.job_queue:
        app.job_queue.run_monthly(
            kirim_rekap_bulanan,
            when=dtime(hour=7, minute=0, tzinfo=WIB),
            day=1,
            name="rekap_bulanan"
        )
        logger.info("Jadwal rekap bulanan terdaftar (setiap tgl 1 jam 07:00 WIB).")
    else:
        logger.warning("JobQueue tidak tersedia — rekap otomatis tidak aktif.")

    logger.info("Bot berjalan...")
    app.run_polling()


if __name__ == "__main__":
    main()
